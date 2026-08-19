#!/usr/bin/env python3
"""Build the direct-workbook audit ledger for Massachusetts Figure 3.

This deliberately does not read ``CurrComp Master.xlsx``.  It applies the
repository README's rule to the lowest-level deposited workbooks:

    credits in Column H on gray pathway rows / cleaned AS Column-H total

Blue rows are unrestricted-elective-only credit and are excluded.  The ratio
is not capped at 100%.  Three visibly duplicated archive records/blocks are
removed by explicit coordinate, with an assertion that every removed pathway
row duplicates an earlier course row in that same sheet.

Run:
    pmt-env/bin/python server/scripts/ma/figure3GrayDetail.py
    pmt-env/bin/python server/scripts/ma/figure3GrayDetail.py --check

Requires openpyxl (the same XLSX reader used by the MA conversion scripts).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


SERVER = Path(__file__).resolve().parents[2]
DATA = SERVER / "data" / "ma"
RECOVERED = DATA / "recovered"
OUTPUT = DATA / "figure3-gray-detail.json"
PDF_FIGURES = DATA / "pdf-figures.json"

UNIVERSITIES = [
    "Bridgewater",
    "Fitchburg",
    "Framingham",
    "MCLA",
    "Salem",
    "UMass Amherst",
    "UMass Boston",
    "UMass Dartmouth",
    "UMass Lowell",
    "Westfield",
    "Worcester",
]

CC_ALIASES = {
    "Mass Bay": "MassBay",
    "Springfield": "Springfield Technical",
    "Wachusett": "Mount Wachusett",
    "M. Wachusett": "Mount Wachusett",
    "Mt. Wachusett": "Mount Wachusett",
}

# These are archive defects, not discretionary course-classification choices.
# Bristol row 28 repeats row 20 verbatim and has no Course ID.  The two
# Roxbury pathway sheets contain a second, no-ID copy of an earlier block.
OBJECTIVE_DUPLICATE_ROWS = {
    ("All CC AS.xlsx", "Bristol"): [28],
    ("All Pathways/Bridgewater.xlsx", "Roxbury"): [66, 67, 72, 73, 75, 76, 77, 81],
    ("All Pathways/UMass Lowell.xlsx", "Roxbury"): [57, 58, 59, 61, 62, 64, 65, 66, 71, 72],
}

GRAY_RGB = {"CCCCCC", "B7B7B7"}
BLUE_RGB = {"A4C2F4"}


def canonical_college(name: str) -> str:
    clean = str(name).strip()
    return CC_ALIASES.get(clean, clean)


def rgb(cell) -> str | None:
    color = cell.fill.fgColor
    if color.type != "rgb" or not isinstance(color.rgb, str):
        return None
    return color.rgb.upper()[-6:]


def row_signature(sheet, row: int) -> tuple[str, str, str, float]:
    values = [str(sheet.cell(row, col).value or "").strip().lower() for col in (2, 3, 4)]
    return (*values, float(sheet.cell(row, 8).value))


def numeric_credit(sheet, row: int) -> float | None:
    value = sheet.cell(row, 8).value
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return float(value)


def duplicate_rows_for(relative_workbook: str, sheet_name: str) -> set[int]:
    return set(OBJECTIVE_DUPLICATE_ROWS.get((relative_workbook, sheet_name), []))


def assert_duplicate_rows(sheet, rows: set[int]) -> None:
    for row in sorted(rows):
        assert sheet.cell(row, 1).value is None, f"{sheet.title}!A{row} must remain blank"
        signature = row_signature(sheet, row)
        earlier = [
            candidate
            for candidate in range(8, row)
            if numeric_credit(sheet, candidate) is not None
            and row_signature(sheet, candidate) == signature
        ]
        assert earlier, f"{sheet.title}!H{row} is no longer an objective earlier-row duplicate"


def cleaned_as_totals() -> tuple[dict[str, float], list[dict]]:
    relative = "All CC AS.xlsx"
    workbook = load_workbook(RECOVERED / relative, data_only=True)
    totals: dict[str, float] = {}
    removals: list[dict] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        removed = duplicate_rows_for(relative, sheet_name)
        assert_duplicate_rows(sheet, removed)
        total = 0.0
        for row in range(8, sheet.max_row + 1):
            credit = numeric_credit(sheet, row)
            if credit is None:
                continue
            if row in removed:
                removals.append({
                    "workbook": relative,
                    "sheet": sheet_name,
                    "row": row,
                    "credit_units": credit,
                    "reason": "No-ID exact duplicate of an earlier AS course row",
                })
                continue
            total += credit
        totals[canonical_college(sheet_name)] = total
    return totals, removals


def selected_pair_sheets(workbook) -> dict[str, str]:
    selected: dict[str, str] = {}
    is_real: dict[str, bool] = {}
    for sheet_name in workbook.sheetnames[1:]:
        if sheet_name == "no clue lol" or sheet_name.upper().startswith("(FAKE)"):
            continue
        real = bool(re.match(r"^\(?REAL\)?\s+", sheet_name, flags=re.IGNORECASE))
        plain = re.sub(r"^\(?REAL\)?\s+", "", sheet_name, flags=re.IGNORECASE).strip()
        college = canonical_college(plain)
        if college in selected and is_real[college] and not real:
            continue
        selected[college] = sheet_name
        is_real[college] = real
    return selected


def workbook_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def build() -> dict:
    pdf = json.loads(PDF_FIGURES.read_text())["fig3_pct_as"]
    as_totals, as_removals = cleaned_as_totals()
    cells: list[dict] = []
    pathway_removals: list[dict] = []
    fill_exceptions: list[dict] = []
    input_files = [RECOVERED / "All CC AS.xlsx"]

    for university in UNIVERSITIES:
        workbook_path = RECOVERED / "All Pathways" / f"{university}.xlsx"
        input_files.append(workbook_path)
        workbook = load_workbook(workbook_path, data_only=True)
        relative = f"All Pathways/{university}.xlsx"
        for college, sheet_name in selected_pair_sheets(workbook).items():
            sheet = workbook[sheet_name]
            removed = duplicate_rows_for(relative, sheet_name)
            assert_duplicate_rows(sheet, removed)
            gray_units = 0.0
            blue_units = 0.0
            gray_rows: list[str] = []
            blue_rows: list[str] = []
            outside_h_rows: list[str] = []

            for row in range(8, sheet.max_row + 1):
                credit = numeric_credit(sheet, row)
                if credit is None:
                    continue
                if row in removed:
                    pathway_removals.append({
                        "workbook": relative,
                        "sheet": sheet_name,
                        "row": row,
                        "credit_units": credit,
                        "reason": "No-ID exact duplicate of an earlier pathway course row",
                    })
                    continue

                row_colors = {rgb(sheet.cell(row, col)) for col in range(2, 9)}
                h_color = rgb(sheet.cell(row, 8))
                if row_colors & GRAY_RGB:
                    gray_units += credit
                    gray_rows.append(f"H{row}")
                    if h_color not in GRAY_RGB:
                        outside_h_rows.append(f"H{row}")
                elif row_colors & BLUE_RGB:
                    blue_units += credit
                    blue_rows.append(f"H{row}")

            if outside_h_rows:
                fill_exceptions.append({
                    "workbook": relative,
                    "sheet": sheet_name,
                    "credit_cells": outside_h_rows,
                    "reason": "The row is visibly gray in B:E but its numeric H cell lost the fill; the row-level gray marker is normalized.",
                })

            denominator = as_totals[college]
            raw_pct = 100 * gray_units / denominator
            display_pct = round(raw_pct)
            final_pct = pdf["cells"][college][university]
            cells.append({
                "pair": f"{university} × {college}",
                "school": university,
                "college": college,
                "archive_gray_units": gray_units,
                "archive_blue_only_units_excluded": blue_units,
                "archive_as_total_units": denominator,
                "archive_gray_detail_pct": round(raw_pct, 10),
                "archive_gray_detail_display_pct": display_pct,
                "final_pdf_pct": final_pct,
                "delta_archive_minus_pdf_pp": round(raw_pct - final_pct, 10),
                "matches_final_pdf_at_printed_precision": display_pct == final_pct,
                "source_workbook": relative,
                "source_sheet": sheet_name,
                "gray_credit_cells": gray_rows,
                "blue_credit_cells_excluded": blue_rows,
                "removed_duplicate_credit_cells": [f"H{row}" for row in sorted(removed)],
            })

    cells.sort(key=lambda cell: (cell["school"], cell["college"]))
    assert len(cells) == 61, f"expected 61 final-PDF pathways, got {len(cells)}"
    assert len({cell["pair"] for cell in cells}) == 61
    raw_mean = sum(cell["archive_gray_detail_pct"] for cell in cells) / len(cells)
    pdf_mean = sum(cell["final_pdf_pct"] for cell in cells) / len(cells)
    matches = sum(cell["matches_final_pdf_at_printed_precision"] for cell in cells)
    assert matches == 42, f"expected 42/61 printed matches, got {matches}"
    assert abs(raw_mean - 64.6824664734) < 1e-8, raw_mean

    all_removals = as_removals + pathway_removals
    duplicate_blocks = []
    for (workbook, sheet), rows in OBJECTIVE_DUPLICATE_ROWS.items():
        duplicate_blocks.append({
            "workbook": workbook,
            "sheet": sheet,
            "rows": rows,
            "record_type": "AS course" if workbook == "All CC AS.xlsx" else "pathway course block",
        })

    return {
        "schema_version": 1,
        "generated_by": "server/scripts/ma/figure3GrayDetail.py",
        "artifact_scope": "Direct rerun of final-PDF Figure 3 from the lowest-level deposited 2024 AS and pathway workbooks; CurrComp Master.xlsx is not read.",
        "formula": {
            "numerator": "Sum Column H credits for rows marked gray (replacement of a bachelor requirement).",
            "denominator": "Sum Column H credits in that community college's AS sheet after removal of one objective trailing duplicate.",
            "blue_rows": "Excluded: blue means unrestricted-general-elective-only credit toward 120 hours.",
            "cap": "None. A raw numerator above the AS total remains above 100%.",
            "display_comparison": "Python/Excel half-even whole-percent rounding compared with the final PDF's printed integer cell.",
            "weighting": "Each of the 61 studied college × university pathways receives equal weight.",
        },
        "format_normalizations": {
            "gray_rgb_values": sorted(GRAY_RGB),
            "blue_rgb_values": sorted(BLUE_RGB),
            "row_fill_exception_count": len(fill_exceptions),
            "row_fill_exceptions": fill_exceptions,
            "note": "Column H normally carries the fill. MCLA–Berkshire marks gray across B:E but leaves H unfilled; this visible row marker is normalized without changing any credit value.",
        },
        "objective_duplicate_blocks_removed": duplicate_blocks,
        "objective_duplicate_rows_removed": all_removals,
        "summary": {
            "cells": len(cells),
            "matches_final_pdf_at_printed_precision": matches,
            "mismatches_final_pdf_at_printed_precision": len(cells) - matches,
            "archive_gray_detail_mean_pct": round(raw_mean, 10),
            "final_pdf_mean_pct": round(pdf_mean, 10),
            "mean_delta_archive_minus_pdf_pp": round(raw_mean - pdf_mean, 10),
        },
        "inputs": [
            {
                "path": str(path.relative_to(SERVER)),
                "sha256": workbook_hash(path),
            }
            for path in input_files
        ] + [{
            "path": str(PDF_FIGURES.relative_to(SERVER)),
            "sha256": workbook_hash(PDF_FIGURES),
            "role": "literal final-PDF Figure 3 transcription used only for the comparison fields",
        }],
        "cells": cells,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true", help="fail if the committed artifact differs")
    args = parser.parse_args()
    rendered = json.dumps(build(), indent=2, ensure_ascii=False) + "\n"
    if args.check:
        if not OUTPUT.exists() or OUTPUT.read_text() != rendered:
            print(f"stale artifact: {OUTPUT}", file=sys.stderr)
            return 1
        print(f"Figure 3 gray-detail artifact is current: {OUTPUT}")
        return 0
    OUTPUT.write_text(rendered)
    print(f"wrote {OUTPUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
