#!/usr/bin/env python
"""Re-run the paper's own math from its own artifacts, faithfully.

The notebooks in the paper's repository do almost no computation:

  - heatmap.ipynb reads the workbook's stored `Lower` / `Upper` columns and
    plots them. The ratios are Excel formulas IN the workbook:
    Lower = COUNTIF(<lower cols>, TRUE) / COUNTA(<lower cols>)
    Upper = COUNTIF(<all cols>,  TRUE) / COUNTA(<all cols>)   <- cumulative
  - course_distribution.ipynb plots hard-coded per-university arrays
    ("Sorry for hard coding these :(") -- no typing logic exists in the repo.
  - ch_currcomp.ipynb reads four hand tabs of CurrComp Master.xlsx
    (% Credit Hours / Credit Hours / Curricular Complexity / Cost) and plots
    boxplots; the only math is seaborn's summary statistics.
  - masstransfer.ipynb reads the MT column.

This script therefore re-executes the real computations behind the figures:

  Fig 1  parse the formula ranges (their own lower/upper split), recompute
         COUNTIF/COUNTA from the boolean cells, and read Excel's cached
         values -- three views that should agree with each other and with
         the published heatmap.
  Fig 2  recompute per-university, per-type articulation shares from the
         matrix under an explicit typing rule, for comparison against the
         notebook's hard-coded arrays.
  Figs 3-6  read the four CurrComp tabs exactly as make_df does and compute
         the summary statistics the paper quotes.

Run: pmt-env/bin/python server/scripts/ma/theirMath.py
Writes: server/data/ma/their-math.json
"""
import json
import re
import statistics
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
RECOVERED = ROOT / 'data' / 'ma' / 'recovered'
OUT = ROOT / 'data' / 'ma' / 'their-math.json'

UNI_LIST = ['Bridgewater', 'Fitchburg', 'Framingham', 'MCLA', 'Salem',
            'UMass Amherst', 'UMass Boston', 'UMass Lowell', 'UMass Dartmouth',
            'Westfield', 'Worcester']

# The notebook's hard-coded Figure 2 inputs, verbatim (course_distribution.ipynb).
HARDCODED_FIG2 = {
    'Computing':  [0.72, 0.88, 0.15, 0.19, 0.30, 0.40, 0.32, 0.66, 0.61, 0.16, 0.25],
    'Math':       [0.65, 0.96, 0.17, 0.40, 0.82, 0.98, 0.67, 0.52, 0.87, 0.84, 1.00],
    'Science':    [1, 1, 1, 0.93, 1, 1, 0.53, 0.78, 0.97, 1, 1],
    'Humanities': [None, None, None, 0.67, None, None, None, 1, 1, 1, 0.47],
}

COL_RE = re.compile(r'([A-Z]+)(\d+)')


def col_to_index(col):
    total = 0
    for ch in col:
        total = total * 26 + (ord(ch) - ord('A') + 1)
    return total


def parse_countif_range(formula):
    """E2:O2 out of =COUNTIF(E2:O2, TRUE) / COUNTA(E2:O2) -> (5, 15)."""
    m = re.search(r'COUNTIF\(([A-Z]+)\d+:([A-Z]+)\d+', formula or '')
    if not m:
        return None
    return col_to_index(m.group(1)), col_to_index(m.group(2))


# Mirror of services/courseTypes.js so the matrix recomputation carries the
# ENGINE's typing (the paper's repo has no typing code of its own). Where the
# hard-coded notebook arrays disagree with this recomputation, the deviation
# is in the paper's hand typing, not in the engine.
PREFIX_TYPES = [
    ({'CS', 'CSE', 'COMP', 'CSCI', 'CSC', 'CIS', 'CAIS', 'CICS', 'COMPSCI'}, 'Computing'),
    ({'MATH', 'MAT', 'MTH', 'MA', 'STAT'}, 'Math'),
    ({'PHYS', 'PHY', 'PHS', 'PHYSIC', 'CHEM', 'BIO', 'BIOL', 'EGR', 'EECE', 'ENGR'}, 'Science'),
]


def course_type(header):
    h = header.lower()
    # The engine's documented exception: discrete math by any name is math.
    if 'discrete' in h:
        return 'Math'
    code = re.search(r'\(([A-Z]+)\s*\d', header)
    if code:
        for prefixes, tname in PREFIX_TYPES:
            if code.group(1) in prefixes:
                return tname
    if re.search(r'math|calculus|statistic|linear algebra|probability', h):
        return 'Math'
    if re.search(r'science|physics|chemistry|biology|lab', h):
        return 'Science'
    if re.search(r'\bcomp|comput|software|programming|data structures|operating systems|networks|algorithm|database|capstone|senior design|upper level elective', h):
        return 'Computing'
    return 'Humanities'


def fig1_and_matrix():
    formulas = openpyxl.load_workbook(RECOVERED / 'Mass Heatmap.xlsx', data_only=False)
    values = openpyxl.load_workbook(RECOVERED / 'Mass Heatmap.xlsx', data_only=True)
    universities = []
    for uni in UNI_LIST:
        wsf, wsv = formulas[uni], values[uni]
        headers = [c.value for c in wsf[1]]
        course_cols = [i for i, h in enumerate(headers, start=1) if i >= 5 and h not in (None, '')]
        lower_range = None
        rows = []
        for rf, rv in zip(wsf.iter_rows(min_row=2), wsv.iter_rows(min_row=2)):
            cc = rf[0].value
            # 'Total' is the tab's own sum row; 'Source' (Salem) is a
            # citation row that carries no verdicts.
            if cc in (None, '', 'Total', 'Source'):
                continue
            lower_formula = rf[1].value if len(rf) > 1 else None
            if isinstance(lower_formula, str) and lower_formula.startswith('='):
                lower_range = parse_countif_range(lower_formula) or lower_range
            cells = {}
            for ci in course_cols:
                cells[ci] = rv[ci - 1].value if ci - 1 < len(rv) else None
            bools = [v for v in cells.values() if isinstance(v, bool)]
            blanks = sum(1 for v in cells.values() if v is None)
            stored_lower = rv[1].value if len(rv) > 1 else None
            stored_all = rv[2].value if len(rv) > 2 else None
            rows.append({
                'cc': cc, 'cells': cells, 'blanks': blanks,
                'stored_lower': stored_lower, 'stored_all': stored_all,
                'mt': rv[3].value if len(rv) > 3 else None,
                'n_bools': len(bools),
            })
        # Recompute their formulas from the raw booleans over the parsed range.
        lo, hi = lower_range if lower_range else (5, 5)
        for row in rows:
            lower_cells = [v for ci, v in row['cells'].items() if lo <= ci <= hi and v is not None]
            all_cells = [v for v in row['cells'].values() if v is not None]
            row['recomputed_lower'] = (sum(1 for v in lower_cells if v is True) / len(lower_cells)) if lower_cells else None
            row['recomputed_all'] = (sum(1 for v in all_cells if v is True) / len(all_cells)) if all_cells else None
            row.pop('cells')
        universities.append({
            'name': uni,
            'headers': [h for i, h in enumerate(headers, start=1) if i in course_cols],
            'lower_formula_range': [lo, hi],
            'lower_course_count': hi - 5 + 1,
            'course_count': len(course_cols),
            'rows': rows,
        })
    return universities


def fig2_recompute(fig1_universities):
    values = openpyxl.load_workbook(RECOVERED / 'Mass Heatmap.xlsx', data_only=True)
    lower_hi = {u['name']: u['lower_formula_range'][1] for u in fig1_universities}
    result = {'typing_rule': 'keyword prefix rule in theirMath.py (their repo has none)', 'universities': []}
    for uni in UNI_LIST:
        ws = values[uni]
        headers = [c.value for c in ws[1]]
        course_cols = [(i, h) for i, h in enumerate(headers, start=1) if i >= 5 and h not in (None, '')]
        by_type, by_type_lower = {}, {}
        for i, header in course_cols:
            t = course_type(header)
            by_type.setdefault(t, []).append(i)
            if i <= lower_hi[uni]:
                by_type_lower.setdefault(t, []).append(i)

        def shares_over(cols_by_type):
            shares = {}
            for tname, cols in cols_by_type.items():
                true_count = 0
                filled = 0
                for row in ws.iter_rows(min_row=2):
                    if row[0].value in (None, '', 'Total', 'Source'):
                        continue
                    for ci in cols:
                        v = row[ci - 1].value if ci - 1 < len(row) else None
                        if isinstance(v, bool):
                            filled += 1
                            if v:
                                true_count += 1
                shares[tname] = round(true_count / filled, 4) if filled else None
            return shares

        result['universities'].append({
            'name': uni,
            'type_columns': {t: len(c) for t, c in by_type.items()},
            'type_columns_lower': {t: len(c) for t, c in by_type_lower.items()},
            'share_of_cells_articulating': shares_over(by_type),
            # Hypothesis under test: the notebook's hard-coded arrays are
            # LOWER-DIVISION shares (the columns their Lower formula spans).
            'share_of_lower_cells_articulating': shares_over(by_type_lower),
        })
    result['hardcoded_in_notebook'] = HARDCODED_FIG2
    return result


def currcomp_tabs():
    wb = openpyxl.load_workbook(RECOVERED / 'CurrComp Master.xlsx', data_only=True)
    tabs = {'% Credit Hours': 'pct_as', 'Credit Hours': 'credit_hours',
            'Curricular Complexity': 'complexity', 'Cost': 'cost'}
    out = {}
    for sheet, key in tabs.items():
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        uni_cols = {h: i for i, h in enumerate(headers, start=1) if h in UNI_LIST}
        cells, resident = {}, {}
        # make_df takes the first 16 data rows: 15 community colleges + Resident.
        for row in ws.iter_rows(min_row=2, max_row=17):
            label = row[0].value
            if label in (None, ''):
                continue
            for uni, ci in uni_cols.items():
                v = row[ci - 1].value
                if not isinstance(v, (int, float)):
                    continue
                if str(label).strip().lower() == 'resident':
                    resident[uni] = v
                else:
                    cells.setdefault(uni, {})[label] = v
        flat = [v for by_cc in cells.values() for v in by_cc.values()]
        # Per-university aggregates, and per-pair deltas against the same
        # university's resident value — candidate sources for the paper's
        # headline numbers (68%, +13, $7,129, +15), identified in the audit.
        uni_means = [statistics.mean(by_cc.values()) for by_cc in cells.values() if by_cc]
        deltas = [v - resident[uni] for uni, by_cc in cells.items() if uni in resident for v in by_cc.values()]
        out[key] = {
            'cells': cells,
            'resident': resident,
            'stats': {
                'n_cells': len(flat),
                'mean_cells': round(statistics.mean(flat), 4) if flat else None,
                'median_cells': round(statistics.median(flat), 4) if flat else None,
                'mean_with_resident': round(statistics.mean(flat + list(resident.values())), 4) if flat else None,
                'mean_of_university_means': round(statistics.mean(uni_means), 4) if uni_means else None,
                'median_of_university_means': round(statistics.median(uni_means), 4) if uni_means else None,
                'mean_delta_vs_resident': round(statistics.mean(deltas), 4) if deltas else None,
                'median_delta_vs_resident': round(statistics.median(deltas), 4) if deltas else None,
            },
        }
    return out


def main():
    universities = fig1_and_matrix()
    fig2 = fig2_recompute(universities)
    all_stored = [r['stored_all'] for u in universities for r in u['rows'] if isinstance(r['stored_all'], (int, float))]
    all_recomputed = [r['recomputed_all'] for u in universities for r in u['rows'] if r['recomputed_all'] is not None]
    exact = sum(
        1 for u in universities for r in u['rows']
        if isinstance(r['stored_all'], (int, float)) and r['recomputed_all'] is not None
        and abs(r['stored_all'] - r['recomputed_all']) < 1e-9
    )
    report = {
        'fig1': {
            'formula': 'Lower =COUNTIF(<lower cols>,TRUE)/COUNTA(<lower cols>); Upper (=ALL levels) over every course column',
            'universities': universities,
            'summary': {
                'cells': len(all_stored),
                'recomputed_exactly': exact,
                'mean_stored_all': round(statistics.mean(all_stored), 4),
                'mean_recomputed_all': round(statistics.mean(all_recomputed), 4),
            },
        },
        'fig2': fig2,
        'currcomp': currcomp_tabs(),
    }
    OUT.write_text(json.dumps(report, indent=1))
    s = report['fig1']['summary']
    print(f"Fig 1: {s['recomputed_exactly']}/{s['cells']} cells reproduce their formula exactly; "
          f"mean stored {s['mean_stored_all']*100:.1f}% vs recomputed {s['mean_recomputed_all']*100:.1f}%")
    cc = report['currcomp']
    print(f"Fig 3 tab mean {cc['pct_as']['stats']['mean_cells']*100:.1f}% | "
          f"Fig 4 tab mean {cc['credit_hours']['stats']['mean_cells']:.1f}h | "
          f"Fig 5 tab mean ${cc['cost']['stats']['mean_cells']:.0f} | "
          f"Fig 6 tab mean {cc['complexity']['stats']['mean_cells']:.1f}")


if __name__ == '__main__':
    main()
